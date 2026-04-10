from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# ═══════════════════════════════════════════════════════════════════════
# ОБЩИЕ СТИЛИ
# ═══════════════════════════════════════════════════════════════════════
def side(s="thin"):
    return Side(border_style=s, color="000000")

B_THIN = Border(left=side(), right=side(), top=side(), bottom=side())
B_MED  = Border(left=side("medium"), right=side("medium"),
                top=side("medium"), bottom=side("medium"))

def fill(hex_):
    return PatternFill("solid", fgColor=hex_)

RUB = '#,##0.00 ₽'
PCT = '0.0%'

def style(cell, font=None, bg=None, align="center", valign="center",
          border=B_THIN, fmt=None, wrap=False):
    if font:   cell.font   = font
    if bg:     cell.fill   = fill(bg)
    if border: cell.border = border
    cell.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
    if fmt:    cell.number_format = fmt

BOLD   = Font(bold=True, size=11)
BOLD12 = Font(bold=True, size=12)
BOLD14 = Font(bold=True, size=14, color="FFFFFF")
ITALIC = Font(italic=True, size=10, color="595959")
WHITE  = Font(bold=True, size=10, color="FFFFFF")
MONO   = Font(name="Consolas", size=10)

MONTHS = ["Янв", "Фев", "Мар", "Апр", "Май", "Июн",
          "Июл", "Авг", "Сен", "Окт", "Ноя", "Дек"]

INCOME_ITEMS = [
    ("Зарплата (основная)",          "💼"),
    ("Подработка / фриланс",         "💻"),
    ("Премия / бонус",               "🎯"),
    ("Социальные выплаты",           "🏛️"),
    ("Проценты / инвестиции",        "📈"),
    ("Подарки / прочие доходы",      "🎁"),
]

EXPENSE_GROUPS = [
    ("ЖИЛЬЁ", [
        "Аренда / ипотека",
        "Коммунальные услуги",
        "Интернет и связь",
        "Ремонт / мебель",
    ]),
    ("ПИТАНИЕ", [
        "Супермаркет / гипермаркет",   # весь чек одной строкой
        "Продукты (малые покупки)",
        "Кафе и рестораны",
    ]),
    ("ТРАНСПОРТ", [
        "Общественный транспорт",
        "Такси",
        "Авто (топливо, парковка)",
    ]),
    ("ЗДОРОВЬЕ", [
        "Медицина / аптека",
        "Спорт и фитнес",
    ]),
    ("ЛИЧНОЕ", [
        "Одежда и обувь",
        "Косметика / гигиена",
        "Образование / курсы",
    ]),
    ("ДОСУГ", [
        "Развлечения / хобби",
        "Подписки (стриминг и др.)",
        "Путешествия / отдых",
    ]),
    ("ФИНАНСЫ", [
        "Кредит / долг",
        "Накопления / инвестиции",
        "Прочие расходы",
    ]),
]

ALL_EXPENSE_ITEMS = [item for _, items in EXPENSE_GROUPS for item in items]
ALL_CATEGORIES    = [i[0] for i in INCOME_ITEMS] + ALL_EXPENSE_ITEMS

# ═══════════════════════════════════════════════════════════════════════
# ЛИСТ 1 — СВОДНАЯ ТАБЛИЦА
# ═══════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Сводка по месяцам"

# ── Заголовок ──────────────────────────────────────────────────────────
TOTAL_COLS = 16  # A..P  (B=категория, C..N=месяцы, O=итого, P=доля)
ws.merge_cells("A1:P1")
ws["A1"] = "ЛИЧНЫЙ БЮДЖЕТ 2026"
ws["A1"].font      = BOLD14
ws["A1"].fill      = fill("1F4E79")
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 38

ws.merge_cells("A2:P2")
ws["A2"] = "Заполняйте ячейки с белым фоном · Формулы пересчитываются автоматически"
ws["A2"].font      = ITALIC
ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 18

# ── Шапка ─────────────────────────────────────────────────────────────
ws.merge_cells("A3:A4"); ws["A3"] = "№"
ws.merge_cells("B3:B4"); ws["B3"] = "Статья"
for i, m in enumerate(MONTHS, start=3):
    col = get_column_letter(i)
    ws.merge_cells(f"{col}3:{col}4"); ws[f"{col}3"] = m
ws.merge_cells("O3:O4"); ws["O3"] = "ИТОГО"
ws.merge_cells("P3:P4"); ws["P3"] = "Доля"

for col in range(1, 17):
    c = ws.cell(3, col)
    c.font      = Font(bold=True, size=10, color="FFFFFF")
    c.fill      = fill("2E75B6")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = B_THIN

ws.row_dimensions[3].height = 26
ws.row_dimensions[4].height = 4

# ── Вспомогательные функции ────────────────────────────────────────────
def sec_header(row, label, bg):
    ws.merge_cells(f"A{row}:P{row}")
    c = ws[f"A{row}"]
    c.value     = f"  {label}"
    c.font      = Font(bold=True, size=11, color="FFFFFF")
    c.fill      = fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border    = B_THIN
    ws.row_dimensions[row].height = 20

def data_row(row, num, label, row_fill, is_expense=False):
    ws.cell(row, 1, num).border    = B_THIN
    ws.cell(row, 1).alignment      = Alignment(horizontal="center", vertical="center")
    ws.cell(row, 2, label).fill    = fill(row_fill)
    ws.cell(row, 2).border         = B_THIN
    ws.cell(row, 2).alignment      = Alignment(horizontal="left", vertical="center")
    for col in range(3, 15):
        c = ws.cell(row, col, 0)
        c.number_format = RUB
        c.fill          = fill("FFFFFF")
        c.border        = B_THIN
        c.alignment     = Alignment(horizontal="right", vertical="center")
    # ИТОГО по строке
    tot = ws.cell(row, 15)
    tot.value         = f"=SUM(C{row}:N{row})"
    tot.number_format = RUB
    tot.font          = BOLD
    tot.fill          = fill("FFEB9C")
    tot.border        = B_THIN
    tot.alignment     = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[row].height = 18

def total_row(row, d_start, d_end, label, bg_label, bg_cells, border=B_MED):
    ws.cell(row, 1, "").border      = border
    ws.cell(row, 2, label).font     = BOLD
    ws.cell(row, 2).fill            = fill(bg_label)
    ws.cell(row, 2).border          = border
    ws.cell(row, 2).alignment       = Alignment(horizontal="left", vertical="center")
    for col in range(3, 15):
        cl = get_column_letter(col)
        c  = ws.cell(row, col)
        c.value         = f"=SUM({cl}{d_start}:{cl}{d_end})"
        c.number_format = RUB
        c.font          = BOLD
        c.fill          = fill(bg_cells)
        c.border        = border
        c.alignment     = Alignment(horizontal="right", vertical="center")
    tot = ws.cell(row, 15)
    tot.value         = f"=SUM(C{row}:N{row})"
    tot.number_format = RUB
    tot.font          = Font(bold=True, size=11)
    tot.fill          = fill(bg_cells)
    tot.border        = border
    tot.alignment     = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[row].height = 22

# ── Расставляем строки ────────────────────────────────────────────────
cur = 5  # текущая строка

# — ДОХОДЫ —
sec_header(cur, "ДОХОДЫ", "375623"); INCOME_SEC = cur; cur += 1
inc_data_start = cur
for n, (label, _) in enumerate(INCOME_ITEMS, 1):
    data_row(cur, n, label, "E2EFDA"); cur += 1
inc_data_end = cur - 1
INC_TOTAL = cur
total_row(cur, inc_data_start, inc_data_end, "ИТОГО ДОХОДОВ",
          "A9D18E", "C6EFCE"); cur += 2

# — РАСХОДЫ (по группам) —
sec_header(cur, "РАСХОДЫ", "833333"); EXPENSE_SEC = cur; cur += 1
exp_data_start = cur
exp_num = 1

# для строки ИТОГО РАСХОДОВ нам нужны все строки с данными
all_exp_data_rows = []

for group_label, items in EXPENSE_GROUPS:
    # подзаголовок группы
    ws.merge_cells(f"A{cur}:P{cur}")
    c = ws[f"A{cur}"]
    c.value     = f"  — {group_label} —"
    c.font      = Font(bold=True, size=10, color="FFFFFF")
    c.fill      = fill("C00000")
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border    = B_THIN
    ws.row_dimensions[cur].height = 17
    cur += 1

    g_start = cur
    for label in items:
        data_row(cur, exp_num, label, "FCE4D6", is_expense=True)
        all_exp_data_rows.append(cur)
        exp_num += 1; cur += 1
    g_end = cur - 1

exp_data_end = cur - 1
EXP_TOTAL = cur
total_row(cur, None, None, "ИТОГО РАСХОДОВ", "FF9999", "FFC7CE")
# переопределяем формулы для суммирования только строк с данными
rows_ref = "+".join(f"O{r}" for r in all_exp_data_rows)
for col in range(3, 16):
    cl  = get_column_letter(col)
    ref = "+".join(f"{cl}{r}" for r in all_exp_data_rows)
    c   = ws.cell(EXP_TOTAL, col)
    c.value = f"={ref}"
cur += 2

# — БАЛАНС —
BAL_ROW = cur
ws.merge_cells(f"A{BAL_ROW}:B{BAL_ROW}")
c = ws[f"A{BAL_ROW}"]
c.value     = "БАЛАНС (Доходы − Расходы)"
c.font      = Font(bold=True, size=12, color="FFFFFF")
c.fill      = fill("1F4E79")
c.alignment = Alignment(horizontal="left", vertical="center")
c.border    = B_MED
ws.row_dimensions[BAL_ROW].height = 28

for col in range(3, 15):
    cl  = get_column_letter(col)
    c   = ws.cell(BAL_ROW, col)
    c.value         = f"={cl}{INC_TOTAL}-{cl}{EXP_TOTAL}"
    c.number_format = RUB
    c.font          = Font(bold=True, size=11)
    c.fill          = fill("BDD7EE")
    c.border        = B_MED
    c.alignment     = Alignment(horizontal="right", vertical="center")

tot = ws.cell(BAL_ROW, 15)
tot.value         = f"=SUM(C{BAL_ROW}:N{BAL_ROW})"
tot.number_format = RUB
tot.font          = Font(bold=True, size=12)
tot.fill          = fill("9DC3E6")
tot.border        = B_MED
tot.alignment     = Alignment(horizontal="right", vertical="center")

# Доля накоплений (баланс / доходы)
save = ws.cell(BAL_ROW, 16)
save.value         = f"=IF(O{INC_TOTAL}<>0, O{BAL_ROW}/O{INC_TOTAL}, 0)"
save.number_format = PCT
save.font          = Font(bold=True, size=11)
save.fill          = fill("9DC3E6")
save.border        = B_MED
save.alignment     = Alignment(horizontal="center", vertical="center")

# Доля по строкам (итого строки / итого доходов)
for r in all_exp_data_rows:
    c = ws.cell(r, 16)
    c.value         = f"=IF(O{INC_TOTAL}<>0, O{r}/O{INC_TOTAL}, 0)"
    c.number_format = PCT
    c.fill          = fill("FFF2CC")
    c.border        = B_THIN
    c.alignment     = Alignment(horizontal="center", vertical="center")

for r in range(inc_data_start, inc_data_end + 1):
    c = ws.cell(r, 16)
    c.value         = f"=IF(O{INC_TOTAL}<>0, O{r}/O{INC_TOTAL}, 0)"
    c.number_format = PCT
    c.fill          = fill("E2EFDA")
    c.border        = B_THIN
    c.alignment     = Alignment(horizontal="center", vertical="center")

# ── Ширина столбцов ───────────────────────────────────────────────────
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 28
for col in range(3, 15):
    ws.column_dimensions[get_column_letter(col)].width = 11
ws.column_dimensions["O"].width = 14
ws.column_dimensions["P"].width = 8

ws.freeze_panes = "C5"

# ═══════════════════════════════════════════════════════════════════════
# ЛИСТ 2 — ЖУРНАЛ ОПЕРАЦИЙ
# ═══════════════════════════════════════════════════════════════════════
wj = wb.create_sheet("Журнал операций")

# Заголовок
wj.merge_cells("A1:G1")
wj["A1"] = "ЖУРНАЛ ДОХОДОВ И РАСХОДОВ"
wj["A1"].font      = BOLD14
wj["A1"].fill      = fill("1F4E79")
wj["A1"].alignment = Alignment(horizontal="center", vertical="center")
wj.row_dimensions[1].height = 34

# Подсказка
wj.merge_cells("A2:G2")
wj["A2"] = ("Для чеков супермаркета: Категория = «Супермаркет / гипермаркет»  →  "
            "вся сумма одной строкой, без разбивки позиций")
wj["A2"].font      = ITALIC
wj["A2"].fill      = fill("FFF2CC")
wj["A2"].alignment = Alignment(horizontal="center", vertical="center")
wj.row_dimensions[2].height = 16

# Шапка
jheaders = ["Дата", "Тип", "Категория", "Описание / Магазин",
            "Сумма", "Счёт / кошелёк", "Примечание"]
jwidths  = [12,      10,    28,          32,
            16,      18,    20]
for i, (h, w) in enumerate(zip(jheaders, jwidths), start=1):
    c = wj.cell(3, i, h)
    c.font      = Font(bold=True, size=10, color="FFFFFF")
    c.fill      = fill("2E75B6")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = B_THIN
    wj.column_dimensions[get_column_letter(i)].width = w
wj.row_dimensions[3].height = 22

# Валидация типа
TYPE_CHOICES = '"Доход,Расход"'
dv_type = DataValidation(type="list", formula1=TYPE_CHOICES, allow_blank=True)
dv_type.sqref = "B4:B203"
wj.add_data_validation(dv_type)

# Валидация категорий (выпадающий список)
# Excel ограничивает строку формулы ~255 символами, поэтому список вынесем
# на скрытый лист-справочник
ws_ref = wb.create_sheet("_Справочник")
ws_ref.sheet_state = "hidden"
ws_ref["A1"] = "Категории"
all_cats_full = [i[0] for i in INCOME_ITEMS] + ALL_EXPENSE_ITEMS
for idx, cat in enumerate(all_cats_full, start=2):
    ws_ref[f"A{idx}"] = cat
cat_count = len(all_cats_full) + 1

dv_cat = DataValidation(
    type="list",
    formula1=f"_Справочник!$A$2:$A${cat_count}",
    allow_blank=True
)
dv_cat.sqref = "C4:C203"
wj.add_data_validation(dv_cat)

SUPERMARKET_CAT = "Супермаркет / гипермаркет"
SUPERMARKET_BG  = "FFF2CC"   # желтоватый фон для строк-супермаркетов

for r in range(4, 204):
    row_bg = "F7F7F7" if r % 2 == 0 else "FFFFFF"

    # Дата
    c = wj.cell(r, 1)
    c.number_format = "DD.MM.YYYY"
    c.fill = fill(row_bg); c.border = B_THIN
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Тип
    c = wj.cell(r, 2)
    c.fill = fill(row_bg); c.border = B_THIN
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Категория
    c = wj.cell(r, 3)
    c.fill = fill(row_bg); c.border = B_THIN
    c.alignment = Alignment(horizontal="left", vertical="center")

    # Описание / Магазин
    c = wj.cell(r, 4)
    c.fill = fill(row_bg); c.border = B_THIN
    c.alignment = Alignment(horizontal="left", vertical="center")

    # Сумма
    c = wj.cell(r, 5)
    c.number_format = RUB
    c.fill = fill(row_bg); c.border = B_THIN
    c.alignment = Alignment(horizontal="right", vertical="center")

    # Счёт и примечание
    for col in (6, 7):
        c = wj.cell(r, col)
        c.fill = fill(row_bg); c.border = B_THIN
        c.alignment = Alignment(horizontal="left", vertical="center")

    wj.row_dimensions[r].height = 17

# Условное форматирование: строки с «Супермаркет» — желтый фон
from openpyxl.formatting.rule import FormulaRule
wj.conditional_formatting.add(
    f"A4:G203",
    FormulaRule(
        formula=[f'=$C4="{SUPERMARKET_CAT}"'],
        fill=fill(SUPERMARKET_BG),
        font=Font(bold=True)
    )
)

# Итоговые строки внизу (данные теперь в строках 4..203)
JTOT  = 205
JTOT2 = 206
JTOT3 = 207
JTOT4 = 208   # отдельная строка — только супермаркеты

def jtot_label(row, text, bg, merge_to="D"):
    wj.merge_cells(f"A{row}:{merge_to}{row}")
    c = wj[f"A{row}"]
    c.value     = text
    c.font      = BOLD
    c.fill      = fill(bg)
    c.border    = B_MED
    c.alignment = Alignment(horizontal="right", vertical="center")
    wj.row_dimensions[row].height = 22

def jtot_sum(row, formula, bg):
    c = wj.cell(row, 5)
    c.value         = formula
    c.number_format = RUB
    c.font          = BOLD
    c.fill          = fill(bg)
    c.border        = B_MED
    c.alignment     = Alignment(horizontal="right", vertical="center")

jtot_label(JTOT,  "ИТОГО ДОХОДОВ",               "C6EFCE")
jtot_sum  (JTOT,  '=SUMIF(B4:B203,"Доход",E4:E203)', "C6EFCE")

jtot_label(JTOT2, "ИТОГО РАСХОДОВ",              "FFC7CE")
jtot_sum  (JTOT2, '=SUMIF(B4:B203,"Расход",E4:E203)', "FFC7CE")

jtot_label(JTOT3, "ОСТАТОК",                     "BDD7EE")
c = wj.cell(JTOT3, 5)
c.value         = f"=E{JTOT}-E{JTOT2}"
c.number_format = RUB; c.font = Font(bold=True, size=12)
c.fill = fill("BDD7EE"); c.border = B_MED
c.alignment = Alignment(horizontal="right", vertical="center")
wj.row_dimensions[JTOT3].height = 26

# Отдельный итог по супермаркетам — видно сколько ушло на большие закупки
jtot_label(JTOT4, f'в т.ч. "{SUPERMARKET_CAT}"', "FFF2CC")
jtot_sum  (JTOT4,
           f'=SUMIF(C4:C203,"{SUPERMARKET_CAT}",E4:E203)',
           "FFF2CC")

wj.freeze_panes = "A4"

# ═══════════════════════════════════════════════════════════════════════
# ЛИСТ 3 — РАСПРЕДЕЛЕНИЕ ЧЕКОВ
# ═══════════════════════════════════════════════════════════════════════
# Идея: вводишь дату + магазин + итого чека, затем раскидываешь
# сумму по категориям — не нужно перебивать каждую позицию.
# Контрольная колонка показывает нераспределённый остаток.
# ═══════════════════════════════════════════════════════════════════════
wr = wb.create_sheet("Распределение чеков")

# ── Заголовок ──────────────────────────────────────────────────────────
SPLIT_CATS = [
    "Продукты питания",
    "Товары для дома",
    "Гигиена / косметика",
    "Одежда и обувь",
    "Электроника / техника",
    "Игрушки / детское",
    "Прочее",
]
N_CATS   = len(SPLIT_CATS)
# Столбцы: A=Дата B=Магазин C=Итого чека D..J=категории K=сумма распред L=остаток
COL_TOTAL   = 3          # C
COL_CAT_S   = 4          # D — первая категория
COL_CAT_E   = COL_CAT_S + N_CATS - 1   # J
COL_DISTRIB = COL_CAT_E + 1            # K — сумма распределено
COL_REST    = COL_CAT_E + 2            # L — остаток
LAST_COL    = COL_REST

def cl(n): return get_column_letter(n)

# Шапка заголовка (2 строки)
wr.merge_cells(f"A1:{cl(LAST_COL)}1")
wr["A1"] = "РАСПРЕДЕЛЕНИЕ ЧЕКОВ ПО КАТЕГОРИЯМ"
wr["A1"].font      = BOLD14
wr["A1"].fill      = fill("1F4E79")
wr["A1"].alignment = Alignment(horizontal="center", vertical="center")
wr.row_dimensions[1].height = 36

# Подсказка
wr.merge_cells(f"A2:{cl(LAST_COL)}2")
wr["A2"] = ("Введите: Дату · Название магазина · Итого чека  →  "
            "заполните нужные категории  →  Остаток должен стать 0")
wr["A2"].font      = ITALIC
wr["A2"].fill      = fill("DEEAF1")
wr["A2"].alignment = Alignment(horizontal="center", vertical="center")
wr.row_dimensions[2].height = 18

# ── Шапка таблицы (строки 3–4) ────────────────────────────────────────
# Строка 3: фиксированные заголовки + заголовки категорий
headers_fixed = ["Дата", "Магазин / Магазин", f"ИТОГО\nчека"]
headers_end   = ["Распределено", "ОСТАТОК\n(должен = 0)"]

for i, h in enumerate(headers_fixed, 1):
    c = wr.cell(3, i, h)
    c.font = Font(bold=True, size=10, color="FFFFFF")
    c.fill = fill("2E75B6")
    c.border = B_THIN
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for j, cat in enumerate(SPLIT_CATS):
    col = COL_CAT_S + j
    c   = wr.cell(3, col, cat)
    c.font = Font(bold=True, size=9, color="FFFFFF")
    c.fill = fill("C00000")
    c.border = B_THIN
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

c = wr.cell(3, COL_DISTRIB, "Распределено")
c.font = Font(bold=True, size=10, color="FFFFFF")
c.fill = fill("375623"); c.border = B_THIN
c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

c = wr.cell(3, COL_REST, "ОСТАТОК\n(→ 0)")
c.font = Font(bold=True, size=10, color="FFFFFF")
c.fill = fill("833333"); c.border = B_THIN
c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

wr.row_dimensions[3].height = 36

# ── Строки данных (50 чеков) ───────────────────────────────────────────
DATA_START = 4
for r in range(DATA_START, DATA_START + 50):
    row_bg = "FFFFFF" if r % 2 == 0 else "F7F7F7"

    # Дата
    c = wr.cell(r, 1)
    c.number_format = "DD.MM.YYYY"
    c.fill = fill(row_bg); c.border = B_THIN
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Магазин
    c = wr.cell(r, 2)
    c.fill = fill(row_bg); c.border = B_THIN
    c.alignment = Alignment(horizontal="left", vertical="center")

    # Итого чека — выделено синим
    c = wr.cell(r, COL_TOTAL)
    c.number_format = RUB
    c.fill = fill("DEEAF1"); c.border = B_MED
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.font = Font(bold=True, size=10)

    # Категории
    for j in range(N_CATS):
        c = wr.cell(r, COL_CAT_S + j)
        c.number_format = RUB
        c.fill = fill(row_bg); c.border = B_THIN
        c.alignment = Alignment(horizontal="right", vertical="center")

    # Распределено = SUM категорий
    cat_s = cl(COL_CAT_S); cat_e = cl(COL_CAT_E)
    c = wr.cell(r, COL_DISTRIB)
    c.value         = f"=SUM({cat_s}{r}:{cat_e}{r})"
    c.number_format = RUB
    c.fill = fill("E2EFDA"); c.border = B_THIN
    c.font = Font(bold=True, size=10)
    c.alignment = Alignment(horizontal="right", vertical="center")

    # Остаток = Итого − Распределено (условное форматирование цветом через значение)
    c = wr.cell(r, COL_REST)
    c.value         = f"={cl(COL_TOTAL)}{r}-{cl(COL_DISTRIB)}{r}"
    c.number_format = RUB
    # Если остаток ≠ 0 — будет видно визуально (красный фон через формулу не работает
    # в openpyxl без ConditionalFormatting, поэтому ставим нейтральный желтый)
    c.fill = fill("FFEB9C"); c.border = B_MED
    c.font = Font(bold=True, size=10)
    c.alignment = Alignment(horizontal="right", vertical="center")

    wr.row_dimensions[r].height = 18

# Условное форматирование: остаток = 0 → зелёный, иначе → красный
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles import Color

rest_range = f"{cl(COL_REST)}{DATA_START}:{cl(COL_REST)}{DATA_START+49}"
wr.conditional_formatting.add(
    rest_range,
    CellIsRule(operator="equal", formula=["0"],
               fill=fill("C6EFCE"), font=Font(bold=True, color="375623"))
)
wr.conditional_formatting.add(
    rest_range,
    CellIsRule(operator="greaterThan", formula=["0"],
               fill=fill("FFC7CE"), font=Font(bold=True, color="9C0006"))
)

# ── Строка ИТОГОВ по листу ─────────────────────────────────────────────
TOT_ROW = DATA_START + 51
wr.merge_cells(f"A{TOT_ROW}:B{TOT_ROW}")
c = wr[f"A{TOT_ROW}"]
c.value = "ИТОГО ПО ВСЕМ ЧЕКАМ"
c.font = Font(bold=True, size=11, color="FFFFFF")
c.fill = fill("1F4E79"); c.border = B_MED
c.alignment = Alignment(horizontal="right", vertical="center")
wr.row_dimensions[TOT_ROW].height = 24

# Итого колонка C (сумма всех чеков)
c = wr.cell(TOT_ROW, COL_TOTAL)
c.value         = f"=SUM({cl(COL_TOTAL)}{DATA_START}:{cl(COL_TOTAL)}{DATA_START+49})"
c.number_format = RUB; c.font = Font(bold=True, size=11)
c.fill = fill("BDD7EE"); c.border = B_MED
c.alignment = Alignment(horizontal="right", vertical="center")

# Итого по каждой категории
for j in range(N_CATS):
    col = COL_CAT_S + j
    c   = wr.cell(TOT_ROW, col)
    c.value         = f"=SUM({cl(col)}{DATA_START}:{cl(col)}{DATA_START+49})"
    c.number_format = RUB; c.font = Font(bold=True, size=10)
    c.fill = fill("FFC7CE"); c.border = B_MED
    c.alignment = Alignment(horizontal="right", vertical="center")

c = wr.cell(TOT_ROW, COL_DISTRIB)
c.value         = f"=SUM({cl(COL_DISTRIB)}{DATA_START}:{cl(COL_DISTRIB)}{DATA_START+49})"
c.number_format = RUB; c.font = Font(bold=True, size=11)
c.fill = fill("C6EFCE"); c.border = B_MED
c.alignment = Alignment(horizontal="right", vertical="center")

c = wr.cell(TOT_ROW, COL_REST)
c.value         = f"={cl(COL_TOTAL)}{TOT_ROW}-{cl(COL_DISTRIB)}{TOT_ROW}"
c.number_format = RUB; c.font = Font(bold=True, size=11)
c.fill = fill("FFEB9C"); c.border = B_MED
c.alignment = Alignment(horizontal="right", vertical="center")

# ── Ширина столбцов ───────────────────────────────────────────────────
wr.column_dimensions["A"].width = 12
wr.column_dimensions["B"].width = 22
wr.column_dimensions[cl(COL_TOTAL)].width = 14
for j in range(N_CATS):
    wr.column_dimensions[cl(COL_CAT_S + j)].width = 14
wr.column_dimensions[cl(COL_DISTRIB)].width = 14
wr.column_dimensions[cl(COL_REST)].width   = 14

wr.freeze_panes = "A4"

# ═══════════════════════════════════════════════════════════════════════
# СОХРАНИТЬ
# ═══════════════════════════════════════════════════════════════════════
out = r"c:\Users\Asus\Documents\projects\my-first-projects\Личный_бюджет_2026_v2.xlsx"
wb.save(out)
print(f"Готово: {out}")
